import io
import re
import json
import os
import openpyxl
from openpyxl import Workbook
import xlrd
import streamlit as st

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────
SCRIPT_DIR     = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_FILE = os.path.join(SCRIPT_DIR, "templates.json")

OUT_HEADERS = [
    "Item ID", "Item Name", "Product/Service",
    "Item Type (Buy/Sell/Both)", "Unit of Measurement", "HSN Code",
    "Item Category", "Default Price", "Regular Buying Price",
    "Wholesale Buying Price", "Regular Selling Price", "MRP",
    "Dealer Price", "Distributor Price", "Current Stock",
    "Min Stock Level", "Max Stock Level", "Tax"
]

TARGET_COLS = [
    "Item Name", "Item ID", "HSN Code",
    "Item Category", "Unit of Measurement", "Tax", "Product/Service"
]

# Weighted keywords: (keyword, weight) — higher weight = stronger signal
WEIGHTED_KEYWORDS = {
    "Item Name":            [("name", 3), ("item", 2), ("product", 2), ("description", 1), ("goods", 1)],
    "Item ID":              [("id", 3), ("code", 2), ("sku", 3), ("no.", 1)],
    "HSN Code":             [("hsn", 3), ("sac", 2)],
    "Item Category":        [("category", 3), ("group", 2), ("parent", 2), ("class", 1), ("family", 1)],
    "Unit of Measurement":  [("unit", 3), ("uom", 3), ("measure", 2), ("base", 1)],
    "Tax":                  [("tax", 3), ("gst", 2), ("igst", 2), ("cgst", 2)],
    "Product/Service":      [("service", 2), ("supply", 2)],
}

# System/UI columns to skip when collecting extra columns
SYSTEM_COLS = {
    "edit", "delete", "checkboxvalue", "checkbox", "widget", "inactive",
    "timestamp", "created", "modified", "updated", "sno", "srno",
    "s.no", "sr.no", "s no", "sr no", "serial no", "active", "status",
    "action", "flag", "chk"
}

ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]')

# Values that identify a column as "Item Category" regardless of column name.
# All entries are pre-normalised (lowercase, no spaces/underscores/hyphens).
# Incoming cell values are normalised the same way before comparing.
def _norm(s):
    return re.sub(r'[\s_\-]+', '', s.lower())

CATEGORY_NORM_TOKENS = {
    "rm", "fg", "sfg", "wip",
    "rawmaterial", "rawmaterials",
    "finishedgood", "finishedgoods",
    "finishproduct", "finishedproduct", "finishgoods",
    "semifinished", "semifinishedgood", "semifinishedgoods",
    "workinprogress",
    "consumable", "consumables",
    "trading", "spares", "spareparts",
    "packingmaterial", "packingmaterials",
    "noninventoryitem", "noninventory",
}

# ─────────────────────────────────────────────────────────────────────────────
# Core helpers
# ─────────────────────────────────────────────────────────────────────────────

def clean(v):
    return ILLEGAL_RE.sub('', v).strip() if isinstance(v, str) else v

def load_templates():
    if os.path.exists(TEMPLATES_FILE):
        with open(TEMPLATES_FILE) as f:
            return json.load(f)
    return {}

def save_templates(data):
    with open(TEMPLATES_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def _fmt_cell(val, fmt):
    """
    Apply Excel number format to preserve leading zeros.
    A format like '00000000' means zero-pad the integer to that many digits.
    """
    if not isinstance(val, (int, float)) or not fmt:
        return val
    # Pure zero-padding format: only '0' digits, no decimal or special chars
    if re.match(r'^0+$', fmt):
        return str(int(val)).zfill(len(fmt))
    return val

def read_file(file_bytes):
    """Read all sheets from xls, xlsx, or csv. Returns {sheet_name: [[row_values]]}"""
    is_xls  = file_bytes[:4] == b'\xd0\xcf\x11\xe0'
    is_xlsx = file_bytes[:4] == b'PK\x03\x04'
    sheets = {}
    if not is_xls and not is_xlsx:
        # Treat as CSV
        import csv as _csv
        try:
            text = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = file_bytes.decode('latin-1')
        reader = _csv.reader(io.StringIO(text))
        rows = []
        for row in reader:
            rows.append([cell.strip() if cell.strip() else None for cell in row])
        return {"Sheet1": rows}
    if is_xls:
        wb = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
        xf_list   = wb.xf_list
        fmt_map   = wb.format_map
        for ws in wb.sheets():
            rows = []
            for r in range(ws.nrows):
                row = []
                for c in range(ws.ncols):
                    cell = ws.cell(r, c)
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        row.append(None)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        v = cell.value
                        num = int(v) if v == int(v) else v
                        # Try to apply leading-zero format from xf record
                        try:
                            xf  = xf_list[cell.xf_index]
                            fmt = fmt_map[xf.format_key].format_str
                            num = _fmt_cell(num, fmt)
                        except Exception:
                            pass
                        row.append(num)
                    else:
                        v = str(cell.value).strip()
                        row.append(v or None)
                rows.append(row)
            sheets[ws.name] = rows
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        for ws in wb.worksheets:
            sheet_rows = []
            for row in ws.iter_rows():
                row_data = []
                for cell in row:
                    val = cell.value
                    if val is not None:
                        val = _fmt_cell(val, cell.number_format)
                    row_data.append(val)
                sheet_rows.append(row_data)
            sheets[ws.title] = sheet_rows
    return sheets

def pick_sheet(sheets):
    """Return the sheet with the most populated rows."""
    return max(sheets, key=lambda s: sum(
        1 for r in sheets[s]
        if any(v is not None and str(v).strip() for v in r)
    ))

def pick_item_master_sheet(sheets):
    """
    Return the sheet whose header row contains HSN, UOM/Unit, AND Tax columns.
    Falls back to pick_sheet() if no such sheet is found.
    """
    HSN_KWS = {"hsn", "sac"}
    UOM_KWS = {"unit", "uom", "measure"}
    TAX_KWS = {"tax", "gst", "igst"}
    for name, rows in sheets.items():
        for row in rows[:20]:
            vals = [str(v).strip().lower() for v in row if v is not None and str(v).strip()]
            has_hsn = any(any(kw in v for kw in HSN_KWS) for v in vals)
            has_uom = any(any(kw in v for kw in UOM_KWS) for v in vals)
            has_tax = any(any(kw in v for kw in TAX_KWS) for v in vals)
            if has_hsn and has_uom and has_tax:
                return name
    return pick_sheet(sheets)

def detect_header(rows):
    """
    Scan first 30 rows to find the header row.
    Scores by: number of string cells + keyword hits × 2.
    """
    HWORDS = {
        "name", "item", "product", "hsn", "unit", "uom", "category", "group",
        "code", "type", "tax", "gst", "description", "id", "price", "qty",
        "quantity", "serial", "service", "rate", "no", "sno", "sr", "brand",
        "barcode", "rack", "sac", "measure", "sku"
    }
    best_i, best_s = 0, -1
    for i, row in enumerate(rows[:30]):
        non_null = [v for v in row if v is not None and str(v).strip()]
        if len(non_null) < 2:
            continue
        strings  = sum(1 for v in non_null if isinstance(v, str))
        keywords = sum(1 for v in non_null if any(w in str(v).lower() for w in HWORDS))
        score = strings + keywords * 2
        if score > best_s:
            best_s, best_i = score, i
    return best_i, [str(v).strip() if v is not None else None for v in rows[best_i]]

def is_system_col(h):
    """Return True if column is a UI/system column that should be skipped."""
    if not h:
        return True
    hl = h.strip().lower().rstrip(".")
    if hl in SYSTEM_COLS:
        return True
    if re.match(r'^s\.?r?\.?n\.?o\.?\s*$', hl):   # SNo, SrNo, S.No, Sr.No …
        return True
    return False

def detect_category_by_values(headers, data_rows, skip_cols):
    """
    Scan column values to find a column that contains category-type data
    (RM, FG, SFG, Raw Material, Finished Goods, etc.).

    Returns the column header if found, else None.
    A column qualifies if ≥40% of its unique non-null values match known category tokens.
    """
    col_idx = {h: i for i, h in enumerate(headers) if h}
    best_col, best_score = None, 0

    for h, i in col_idx.items():
        if h in skip_cols or is_system_col(h):
            continue
        unique_vals = set()
        for row in data_rows:
            if i < len(row) and row[i] is not None:
                v = str(row[i]).strip().lower()
                if v:
                    unique_vals.add(v)
        if not unique_vals:
            continue

        # Count how many unique values match category tokens (normalised)
        hits = sum(
            1 for v in unique_vals
            if _norm(v) in CATEGORY_NORM_TOKENS
            or any(tok in _norm(v) for tok in CATEGORY_NORM_TOKENS if len(tok) > 3)
        )
        ratio = hits / len(unique_vals)
        if ratio >= 0.4 and ratio > best_score:
            best_score, best_col = ratio, h

    return best_col


def auto_map(headers, data_rows):
    """
    Map client column headers to target columns automatically.

    Strategy:
      1. Exact name match (case-insensitive)
      2. Category detection by column VALUES (RM / FG / SFG patterns)
      3. Weighted keyword scoring for remaining columns
      4. Item ID fallback: uniqueness check on all unmapped columns

    Returns:
      mapping    – {target_col: client_col_header}
      extra_cols – list of useful unmapped column headers
    """
    result, used = {}, set()

    # ── Pass 1: exact name match ─────────────────────────────────────────
    for h in headers:
        if not h:
            continue
        for t in TARGET_COLS:
            if t not in used and h.lower() == t.lower():
                result[t] = h
                used.add(t)
                break

    # ── Pass 2: category detection by column VALUES ──────────────────────
    if "Item Category" not in result:
        cat_col = detect_category_by_values(headers, data_rows, set(result.values()))
        if cat_col:
            result["Item Category"] = cat_col
            used.add("Item Category")

    # ── Pass 3: weighted keyword scoring ────────────────────────────────
    def _kw_match(kw, h_l):
        # Short keywords (≤2 chars, e.g. "id", "no") need a whole-word match
        # to avoid false hits like "id" inside "widget" or "modified".
        if len(kw) <= 2:
            return bool(re.search(r'\b' + re.escape(kw) + r'\b', h_l))
        return kw in h_l

    mapped_cols  = set(result.values())   # source cols already assigned
    used_sources = set(mapped_cols)        # keep this updated during the loop
    cands = []
    for h in headers:
        if not h or h in mapped_cols or is_system_col(h):
            continue
        h_l = h.lower()
        for t, wkws in WEIGHTED_KEYWORDS.items():
            if t in used:
                continue
            score = sum(w for kw, w in wkws if _kw_match(kw, h_l))
            if score:
                cands.append((score, t, h))

    for score, t, h in sorted(cands, reverse=True):
        if t not in used and h not in used_sources:
            result[t] = h
            used.add(t)
            used_sources.add(h)  # prevent same source col from mapping to two targets

    # ── Pass 4: Item ID by uniqueness (if still not found) ───────────────
    if "Item ID" not in result and data_rows:
        col_idx    = {h: i for i, h in enumerate(headers) if h}
        mapped_set = set(result.values())
        # Keywords that signal "this column holds an item identifier"
        ID_KEYWORDS = {"id", "code", "sku", "no", "number", "num",
                       "ref", "key", "barcode", "upc", "ean"}

        candidates = []
        for h, i in col_idx.items():
            if h in mapped_set or is_system_col(h):
                continue
            vals = [
                str(row[i]).strip() for row in data_rows
                if i < len(row) and row[i] is not None and str(row[i]).strip()
            ]
            if len(vals) < 5:
                continue

            # Skip sequential integers — those are just row/serial numbers
            try:
                nums = sorted(int(v) for v in vals)
                if nums == list(range(nums[0], nums[0] + len(nums))):
                    continue
            except (ValueError, TypeError):
                pass

            ratio = len(set(vals)) / len(vals)
            if ratio < 0.95:
                continue

            # Score by how "ID-like" the column name looks
            h_l      = h.lower()
            id_score = sum(1 for kw in ID_KEYWORDS if kw in h_l)
            candidates.append((id_score, ratio, h))

        if candidates:
            # Best = highest id_name_score first, then highest uniqueness ratio
            candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
            result["Item ID"] = candidates[0][2]

    # ── Extra cols: useful unmapped columns ──────────────────────────────
    mapped_set = set(result.values())
    extra_cols = [
        h for h in headers
        if h and h not in mapped_set and not is_system_col(h)
    ]

    return result, extra_cols

def find_template(headers, templates):
    """Find best matching saved template (≥85% column overlap)."""
    fp = set(h for h in headers if h)
    best_name, best_r = None, 0
    for name, tmpl in templates.items():
        t_fp = set(tmpl.get("fingerprint", []))
        if not t_fp:
            continue
        r = len(fp & t_fp) / max(len(fp), len(t_fp))
        if r > best_r:
            best_r, best_name = r, name
    if best_r >= 0.85:
        return best_name, templates[best_name]
    return None, None

def do_convert(rows, header_idx, mapping, extra_cols):
    """Convert rows to Product_Add format, appending extra columns at end."""
    hdrs = [str(v).strip() if v is not None else None for v in rows[header_idx]]
    col  = {h: i for i, h in enumerate(hdrs) if h}
    tidx = {t: col[c] for t, c in mapping.items() if c and c in col}
    eidx = [(h, col[h]) for h in extra_cols if h in col]

    out = []
    for row in rows[header_idx + 1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def g(t):
            i = tidx.get(t)
            if i is None or i >= len(row):
                return None
            v = row[i]
            return clean(str(v)) if (v is not None and str(v).strip()) else None

        name = g("Item Name")
        if not name:
            continue

        ps    = g("Product/Service")
        ps_out = "Service" if (ps and "service" in ps.lower()) else "Product"

        tax = g("Tax")
        if not tax or tax == "0":
            tax_out = None
        else:
            try:
                tax_out = int(float(tax))
            except (ValueError, TypeError):
                tax_out = None

        row_out = [
            g("Item ID"), name, ps_out, "Both",
            g("Unit of Measurement") or "", g("HSN Code"),
            g("Item Category") or "",
            None, None, None, None, None, None, None,
            None, None, None, tax_out
        ]
        for _, ei in eidx:
            v = row[ei] if ei < len(row) else None
            row_out.append(clean(str(v)) if (v is not None and str(v).strip()) else None)
        out.append(row_out)
    return out

def make_xlsx(rows, extra_col_headers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(OUT_HEADERS + list(extra_col_headers))
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def out_filename(fname):
    m = re.search(r'\(([^)]+)\)', fname)
    cn = m.group(1) if m else fname.rsplit('.', 1)[0]
    return f"Product_Add_({cn}).xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# Network Master helpers
# ─────────────────────────────────────────────────────────────────────────────

NETWORK_OUT_HEADERS = [
    "Company Name", "Buyer/Supplier/Both", "Company Reference ID",
    "TCS Type", "Company Email", "Company Contact Number",
    "Address Line 1", "Address Line 2", "City", "State", "Country",
    "PIN Code", "GSTIN", "GSTIN Type",
    "Contact Person First Name", "Contact Person Last Name", "Contact Person Email"
]

@st.cache_data(show_spinner=False)
def load_pincode_db():
    path = os.path.join(SCRIPT_DIR, "pincode_db.json")
    if os.path.exists(path):
        with open(path) as f:
            return json.load(f)
    return {}

GSTIN_STATE_MAP = {
    "01":"Jammu & Kashmir","02":"Himachal Pradesh","03":"Punjab","04":"Chandigarh",
    "05":"Uttarakhand","06":"Haryana","07":"Delhi","08":"Rajasthan","09":"Uttar Pradesh",
    "10":"Bihar","11":"Sikkim","12":"Arunachal Pradesh","13":"Nagaland","14":"Manipur",
    "15":"Mizoram","16":"Tripura","17":"Meghalaya","18":"Assam","19":"West Bengal",
    "20":"Jharkhand","21":"Odisha","22":"Chhattisgarh","23":"Madhya Pradesh",
    "24":"Gujarat","25":"Daman and Diu","26":"Dadra and Nagar Haveli","27":"Maharashtra",
    "28":"Andhra Pradesh","29":"Karnataka","30":"Goa","31":"Lakshadweep","32":"Kerala",
    "33":"Tamil Nadu","34":"Puducherry","35":"Andaman and Nicobar Islands",
    "36":"Telangana","37":"Andhra Pradesh","38":"Ladakh","97":"Other Territory",
}

def state_from_gstin(gstin):
    if gstin and len(gstin) >= 2:
        return GSTIN_STATE_MAP.get(gstin[:2].zfill(2))
    return None

def clean_gstin(v):
    if not v:
        return None
    s = str(v).strip()
    if s.lower() in ("none", "n/a", "", "0"):
        return None
    return s

def clean_pin(v):
    if not v:
        return None
    digits = re.sub(r"[^\d]", "", str(v).strip())
    return digits[:6] if len(digits) >= 6 else (digits if digits else None)

_NAME_PREFIXES = {
    "mr", "mrs", "ms", "dr", "shri", "smt", "prof",
    "er", "ca", "cs", "adv", "col", "capt", "gen", "lt"
}

def split_name(full):
    if not full:
        return None, None
    # Remove numeric values (phone numbers, IDs embedded in name)
    s = re.sub(r'\b\d+\b', '', full).strip()
    # Normalize spaces
    parts = s.split()
    # Strip honorific prefixes (Mr. / Mrs. / Dr. etc.)
    while parts and re.sub(r'[.\s]', '', parts[0]).lower() in _NAME_PREFIXES:
        parts = parts[1:]
    if not parts:
        return None, None
    if len(parts) == 1:
        return parts[0], None
    return parts[0], " ".join(parts[1:])

def parse_mshriy_address(addr_str):
    """
    Parse a combined address like
    'D 24, SUNPLAZA, VADSAR ROAD, Vadodara, Gujarat, 390010'
    into (addr1, addr2, city, state, pincode).
    Convention: last = pincode, prev = state, prev-prev = city, rest = address.
    """
    if not addr_str:
        return None, None, None, None, None
    parts = [p.strip() for p in addr_str.split(",") if p.strip()]
    pincode = city = state = None

    # Find pincode — last purely-numeric 6-digit segment
    pin_idx = None
    for i in range(len(parts) - 1, -1, -1):
        d = re.sub(r"[^\d]", "", parts[i])
        if len(d) == 6:
            pincode = d
            pin_idx = i
            break

    if pin_idx is not None and pin_idx >= 2:
        state = parts[pin_idx - 1]
        city  = parts[pin_idx - 2]
        addr_parts = parts[: pin_idx - 2]
    elif pin_idx is not None and pin_idx == 1:
        state = parts[0]
        addr_parts = []
    else:
        addr_parts = parts[: pin_idx] if pin_idx else parts

    # Split address into two lines at midpoint
    if not addr_parts:
        addr1 = addr2 = None
    elif len(addr_parts) == 1:
        addr1, addr2 = addr_parts[0], None
    else:
        mid = max(1, len(addr_parts) // 2)
        addr1 = ", ".join(addr_parts[:mid]) or None
        addr2 = ", ".join(addr_parts[mid:]) or None

    return addr1, addr2, city, state, pincode

STATE_ABBR_TO_NAME = {
    "AP": "Andhra Pradesh",   "AR": "Arunachal Pradesh", "AS": "Assam",
    "BR": "Bihar",            "CG": "Chhattisgarh",      "GA": "Goa",
    "GJ": "Gujarat",          "HR": "Haryana",            "HP": "Himachal Pradesh",
    "JH": "Jharkhand",        "KA": "Karnataka",          "KL": "Kerala",
    "MP": "Madhya Pradesh",   "MH": "Maharashtra",        "MN": "Manipur",
    "ML": "Meghalaya",        "MZ": "Mizoram",            "NL": "Nagaland",
    "OD": "Odisha",           "OR": "Odisha",             "PB": "Punjab",
    "RJ": "Rajasthan",        "SK": "Sikkim",             "TN": "Tamil Nadu",
    "TS": "Telangana",        "TR": "Tripura",            "UP": "Uttar Pradesh",
    "UK": "Uttarakhand",      "WB": "West Bengal",
    "AN": "Andaman and Nicobar Islands", "CH": "Chandigarh",
    "DN": "Dadra and Nagar Haveli",      "DD": "Daman and Diu",
    "DL": "Delhi",            "JK": "Jammu & Kashmir",   "LA": "Ladakh",
    "LD": "Lakshadweep",      "PY": "Puducherry",
}

def parse_combined_address(addr_str):
    """
    Parse a combined address like:
    'No 71 1st Cross 2nd Main Road New Tharagupet Bengaluru KA- 560002'
    '311/47-D, Sandaipet Main Road, Shevapet Salem TN- 636002'

    Pattern: ...City STATE_ABBR- PINCODE (last word before state abbr = city).
    Falls back to parse_mshriy_address for comma-separated format.

    Returns (addr1, addr2, city, state, pincode).
    """
    if not addr_str:
        return None, None, None, None, None

    s = addr_str.strip()
    m = re.search(r'\b([A-Z]{2})\s*[-\u2013]\s*(\d{6})\s*$', s)
    if m:
        abbr    = m.group(1)
        pincode = m.group(2)
        state   = STATE_ABBR_TO_NAME.get(abbr)
        prefix  = s[:m.start()].strip().rstrip(',').strip()
        words   = prefix.split()
        if words:
            city          = words[-1]
            addr_str_rest = " ".join(words[:-1])
        else:
            city          = None
            addr_str_rest = ""
        addr_words = addr_str_rest.split()
        if not addr_words:
            addr1, addr2 = None, None
        elif len(addr_words) <= 5:
            addr1, addr2 = addr_str_rest, None
        else:
            mid   = len(addr_words) // 2
            addr1 = " ".join(addr_words[:mid]) or None
            addr2 = " ".join(addr_words[mid:]) or None
        return addr1, addr2, city, state, pincode

    # Fallback: try comma-separated format
    return parse_mshriy_address(addr_str)

def _tally_detect_sheet(sheets):
    """Pick the SVNaturalLanguage sheet from Tally exports, else most populated."""
    if "SVNaturalLanguage" in sheets:
        return "SVNaturalLanguage"
    for name, rows in sheets.items():
        for row in rows[:5]:
            vals = [str(v).lower() for v in row if v]
            if any("$name" in v for v in vals):
                return name
    return pick_sheet(sheets)

ITEM_MASTER_SIGNALS = {
    "hsn", "hsn code", "sac", "item id", "item code", "sku",
    "uom", "unit of measurement", "item category", "item type",
    "mrp", "selling price", "buying price",
    "regular selling price", "regular buying price",
}

def is_item_master_file(sheets):
    """Returns (True, matched_signals) if file looks like an Item Master."""
    for srows in sheets.values():
        for row in srows[:15]:
            vals = {str(v).strip().lower() for v in row if v is not None and str(v).strip()}
            matched = vals & ITEM_MASTER_SIGNALS
            if len(matched) >= 2:
                return True, matched
    return False, set()

def detect_network_format(rows):
    """
    Returns ('tally', header_idx) or ('mshriy', header_idx) or
            ('generic', header_idx) or ('unknown', 0).
    Generic: single combined ADDRESS column + GSTIN + vendor/party name col.
    """
    for i, row in enumerate(rows[:15]):
        vals   = [str(v).strip().lower() for v in row if v is not None and str(v).strip()]
        joined = " ".join(vals)
        if "$name" in joined or "$_primarygroup" in joined:
            return "tally", i
        if "name of ledger" in joined or ("sl" in joined and "name" in joined):
            return "mshriy", i
        # Generic: has a combined address column + GSTIN + party name column
        has_addr  = any(v == "address" for v in vals)
        has_gstin = any("gstin" in v for v in vals)
        has_name  = any(any(kw in v for kw in ("vendor", "party", "supplier", "buyer", "customer")) for v in vals)
        if has_addr and has_gstin and has_name:
            return "generic", i
    return "unknown", 0

def _g(row, col_map, *keys):
    """Get first non-empty value from row by column name keys."""
    for k in keys:
        i = col_map.get(k)
        if i is not None and i < len(row):
            v = row[i]
            if v is not None:
                s = clean(str(v))
                if s and s.lower() not in ("none", "0.00", "0"):
                    return s
    return None

def convert_tally_parties(rows, header_idx):
    hdrs = [str(v).strip() if v is not None else "" for v in rows[header_idx]]
    col  = {h: i for i, h in enumerate(hdrs)}
    parties = []

    for row in rows[header_idx + 1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        name = _g(row, col, "$Name")
        if not name:
            continue

        group   = _g(row, col, "$_PrimaryGroup") or ""
        group_l = group.lower()
        if "sundry debtor" in group_l:
            party_type, is_red = "Buyer", False
        elif "sundry creditor" in group_l:
            party_type, is_red = "Supplier", False
        else:
            party_type, is_red = group, True

        addr1  = _g(row, col, "$_Address1")
        addr2  = _g(row, col, "$_Address2")
        addr3  = _g(row, col, "$_Address3")
        addr2  = (addr2 + ", " + addr3) if addr2 and addr3 else (addr2 or addr3)
        state  = _g(row, col, "$PriorStateName")
        country= _g(row, col, "$CountryName") or "India"
        pin    = clean_pin(_g(row, col, "$pincode", "$Pincode"))
        gstin  = clean_gstin(_g(row, col, "$_PartyGSTIN", "$PartyGSTIN"))
        mobile = _g(row, col, "$LedgerMobile")
        email  = _g(row, col, "$email", "$Email")
        fname, lname = split_name(_g(row, col, "$LedgerContact"))

        # Fill state from GSTIN if still missing
        if not state and gstin:
            state = state_from_gstin(gstin)

        parties.append({
            "Company Name": name,
            "Buyer/Supplier/Both": party_type,
            "Company Reference ID": None, "TCS Type": None,
            "Company Email": email,
            "Company Contact Number": mobile,
            "Address Line 1": addr1, "Address Line 2": addr2,
            "City": None, "State": state, "Country": country,
            "PIN Code": pin,
            "GSTIN": gstin, "GSTIN Type": "Regular" if gstin else None,
            "Contact Person First Name": fname,
            "Contact Person Last Name": lname,
            "Contact Person Email": None,
            "_is_red": is_red,
        })
    return parties

def convert_mshriy_parties(rows, header_idx):
    hdrs = [str(v).strip() if v is not None else "" for v in rows[header_idx]]
    col  = {h: i for i, h in enumerate(hdrs)}
    parties = []

    for row in rows[header_idx + 1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        name = _g(row, col, "Name of Ledger")
        if not name:
            continue

        group   = _g(row, col, "Under") or ""
        group_l = group.lower()
        if "sundry debtor" in group_l:
            party_type, is_red = "Buyer", False
        elif "sundry creditor" in group_l:
            party_type, is_red = "Supplier", False
        else:
            party_type, is_red = group, True

        addr_raw              = _g(row, col, "Address")
        addr1, addr2, city, state_p, pin_p = parse_mshriy_address(addr_raw)
        state  = _g(row, col, "State Name") or state_p
        pin    = clean_pin(_g(row, col, "Pincode", "PIN", "Pin Code")) or clean_pin(pin_p)
        gstin  = clean_gstin(_g(row, col, "GSTIN/UIN", "GSTIN"))
        email  = _g(row, col, "Mail ID", "Email")
        mobile = _g(row, col, "Contact No.", "Mobile", "Phone")

        # Fill state from GSTIN if missing
        if not state and gstin:
            state = state_from_gstin(gstin)

        parties.append({
            "Company Name": name,
            "Buyer/Supplier/Both": party_type,
            "Company Reference ID": None, "TCS Type": None,
            "Company Email": email,
            "Company Contact Number": mobile,
            "Address Line 1": addr1, "Address Line 2": addr2,
            "City": city, "State": state, "Country": "India",
            "PIN Code": pin,
            "GSTIN": gstin, "GSTIN Type": "Regular" if gstin else None,
            "Contact Person First Name": None,
            "Contact Person Last Name": None, "Contact Person Email": None,
            "_is_red": is_red,
        })
    return parties

def convert_generic_parties(rows, header_idx):
    """
    Convert generic combined-address format (e.g. Tranzact):
    VENDOR NAME | ADDRESS (full, combined) | GSTIN
    """
    hdrs = [str(v).strip() if v is not None else "" for v in rows[header_idx]]
    col  = {h: i for i, h in enumerate(hdrs)}

    # Auto-detect column names
    name_col = addr_col = gstin_col = email_col = mobile_col = None
    for h in hdrs:
        hl = h.strip().lower()
        if any(kw in hl for kw in ("vendor", "party", "supplier", "buyer", "customer")):
            if name_col is None:
                name_col = h
        elif "name" in hl and name_col is None:
            name_col = h
        if "address" in hl and addr_col is None:
            addr_col = h
        if "gstin" in hl and gstin_col is None:
            gstin_col = h
        if ("email" in hl or "mail" in hl) and email_col is None:
            email_col = h
        if any(kw in hl for kw in ("mobile", "phone", "contact")) and mobile_col is None:
            mobile_col = h

    parties = []
    for row in rows[header_idx + 1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        name = _g(row, col, name_col) if name_col else None
        if not name:
            continue

        addr_raw = _g(row, col, addr_col) if addr_col else None
        addr1, addr2, city, state, pin_parsed = (
            parse_combined_address(addr_raw) if addr_raw else (None, None, None, None, None)
        )

        gstin  = clean_gstin(_g(row, col, gstin_col)) if gstin_col else None
        pin    = clean_pin(pin_parsed)
        email  = _g(row, col, email_col)  if email_col  else None
        mobile = _g(row, col, mobile_col) if mobile_col else None

        if not state and gstin:
            state = state_from_gstin(gstin)

        parties.append({
            "Company Name": name,
            "Buyer/Supplier/Both": "Both",
            "Company Reference ID": None, "TCS Type": None,
            "Company Email": email,
            "Company Contact Number": mobile,
            "Address Line 1": addr1, "Address Line 2": addr2,
            "City": city, "State": state, "Country": "India",
            "PIN Code": pin,
            "GSTIN": gstin, "GSTIN Type": "Regular" if gstin else None,
            "Contact Person First Name": None,
            "Contact Person Last Name": None, "Contact Person Email": None,
            "_is_red": False,
        })
    return parties

def apply_pincode_lookup(parties, pincode_db):
    for p in parties:
        pin = p.get("PIN Code")
        if pin:
            entry = pincode_db.get(str(pin).zfill(6))
            if entry:
                if not p.get("City"):
                    p["City"] = entry.get("c") or None
                if not p.get("State"):
                    p["State"] = entry.get("s") or None
    return parties

def fill_addr2_with_city(parties):
    """If Address Line 2 is empty, copy City into it."""
    for p in parties:
        if not p.get("Address Line 2") and p.get("City"):
            p["Address Line 2"] = p["City"]
    return parties

def split_network_sheets(parties):
    """
    Sheet 1 (Ready to upload)        : Address Line 1 AND valid 6-digit PIN, non-duplicate GSTIN
    Sheet 2 (Have GSTIN)             : Not ready but GSTIN present
    Sheet 3 (Need to update manually): everything else, plus bad-pin rows (_bad_pin=True)
    Sheet 4 (Duplicate GSTINs)       : rows whose GSTIN already appears in Sheet 1
    """
    ready, have_gstin, manual, duplicates = [], [], [], []
    seen_gstins = set()

    temp_ready = []
    for p in parties:
        pin       = p.get("PIN Code")
        pin_valid = bool(pin and len(str(pin)) >= 6)
        has_addr  = bool(p.get("Address Line 1"))

        if has_addr and pin_valid:
            temp_ready.append(p)
        elif has_addr and pin and not pin_valid:
            p2 = dict(p)
            p2["_bad_pin"] = True
            manual.append(p2)
        elif p.get("GSTIN"):
            have_gstin.append(p)
        else:
            manual.append(p)

    for p in temp_ready:
        gstin = p.get("GSTIN")
        if gstin and gstin in seen_gstins:
            duplicates.append(p)
        else:
            if gstin:
                seen_gstins.add(gstin)
            ready.append(p)

    return ready, have_gstin, manual, duplicates

def make_network_xlsx(ready, have_gstin, manual, duplicates=None):
    from openpyxl.styles import PatternFill
    RED_FILL     = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    PIN_RED_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    PIN_COL_IDX  = NETWORK_OUT_HEADERS.index("PIN Code") + 1   # 1-indexed

    wb = Workbook()
    def write_sheet(ws, parties):
        ws.append(NETWORK_OUT_HEADERS)
        for p in parties:
            ws.append([p.get(h) for h in NETWORK_OUT_HEADERS])
            r = ws.max_row
            if p.get("_is_red"):
                for c in range(1, len(NETWORK_OUT_HEADERS) + 1):
                    ws.cell(row=r, column=c).fill = RED_FILL
            if p.get("_bad_pin"):
                ws.cell(row=r, column=PIN_COL_IDX).fill = PIN_RED_FILL

    ws1 = wb.active;  ws1.title = "Ready to upload";      write_sheet(ws1, ready)
    ws2 = wb.create_sheet("Have GSTIN");                   write_sheet(ws2, have_gstin)
    ws3 = wb.create_sheet("Need to update manually");      write_sheet(ws3, manual)
    ws4 = wb.create_sheet("Duplicate GSTINs");             write_sheet(ws4, duplicates or [])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def net_filename(fname):
    m  = re.search(r"\(([^)]+)\)", fname)
    cn = m.group(1) if m else fname.rsplit(".", 1)[0]
    return f"Network_Add_({cn}).xlsx"

# ─────────────────────────────────────────────────────────────────────────────
# BOM Upload helpers
# ─────────────────────────────────────────────────────────────────────────────

BOM_FG_HEADERS = [
    "Sl_No", "FG Item ID", "FG Item Name", "FG UOM", "BOM Number", "BOM Name",
    "FG Store", "RM Store", "Scrap Store", "BOM Description", "FG Cost Allocation",
    "FG Comment", "Comment"
]

BOM_RM_HEADERS = [
    "Sl_No", "FG Item ID", "BOM Number", "#", "Item Id", "Item Description",
    "Quantity", "Unit", "Comment"
]

BOM_SCRAP_HEADERS = [
    "Sl_No", "FG Item ID", "BOM Number", "#", "Item Id", "Item Description",
    "Quantity", "Unit", "Cost Allocation", "Comment"
]

BOM_ROUTING_HEADERS = [
    "Sl_No", "FG Item ID", "BOM Number", "#", "Routing Number", "Routing Name", "Comment"
]

BOM_OTHER_CHARGES_HEADERS = [
    "Sl_No", "FG Item ID", "BOM Number",
    "Labour Cost", "Labour Comment",
    "Machinery Cost", "Machinery Comment",
    "Electricity Cost", "Electricity Comment",
    "Other Cost", "Other Comment"
]

BOM_INSTRUCTIONS_HEADERS = [
    "Sheet Name", "Field Name", "Data Type", "Mandatory (Yes/No)", "Comment"
]

BOM_INSTRUCTIONS_ROWS = [
    ("FG",  "Serial Number",     "Integer", "Yes", "Use | for alternate item. Eg. 1 | 1"),
    ("FG",  "FG Item ID",        "Text",    "Yes", ""),
    ("FG",  "FG Item Name",      "Text",    "Yes", ""),
    ("FG",  "FG UOM",            "Text",    "Yes", ""),
    ("FG",  "BOM Number",        "Text",    "Yes", "Use BOM Series Name for automatic document series"),
    ("FG",  "BOM Name",          "Text",    "Yes", ""),
    ("FG",  "FG Store",          "Text",    "Yes", ""),
    ("FG",  "RM Store",          "Text",    "Yes", ""),
    ("FG",  "Scrap Store",       "Text",    "Yes", ""),
    ("FG",  "BoM Description",   "Text",    "No",  ""),
    ("FG",  "FG Cost Allocation","Float",   "Yes", ""),
    ("FG",  "FG Comment",        "Text",    "No",  ""),
    ("FG",  "CF1",               "Depends", "Depends", ""),
    ("RM",  "Serial Number",     "Integer", "Yes", ""),
    ("RM",  "FG Item ID",        "Text",    "Yes", ""),
    ("RM",  "BOM Number",        "Text",    "No",  ""),
    ("RM",  "#",                 "Integer", "Yes", "Use | for alternate item.  Eg. 1 | 1"),
    ("RM",  "Item ID",           "Text",    "Yes", ""),
    ("RM",  "Item Description",  "Text",    "Yes", ""),
    ("RM",  "Quantity",          "Float",   "Yes", ""),
    ("RM",  "Unit",              "Text",    "Yes", ""),
    ("RM",  "Comment",           "Text",    "No",  ""),
    ("RM",  "CF1",               "Depends", "Depends", ""),
    ("Scrap", "Serial Number",   "Integer", "Yes", ""),
    ("Scrap", "FG Item ID",      "Text",    "Yes", ""),
    ("Scrap", "BOM Number",      "Text",    "No",  ""),
    ("Scrap", "#",               "Integer", "Yes", ""),
    ("Scrap", "Item ID",         "Text",    "Yes", ""),
    ("Scrap", "Item Description","Text",    "No",  ""),
    ("Scrap", "Quantity",        "Float",   "Yes", ""),
    ("Scrap", "Unit",            "Text",    "Yes", ""),
    ("Scrap", "Cost Allocation", "Float",   "Yes", ""),
    ("Scrap", "Comment",         "Text",    "No",  ""),
    ("Scrap", "CF1",             "Depends", "Depends", ""),
    ("Routing", "Serial Number", "Integer", "Yes", ""),
    ("Routing", "FG Item ID",    "Text",    "Yes", ""),
    ("Routing", "BOM Number",    "Text",    "No",  ""),
    ("Routing", "#",             "Integer", "Yes", ""),
    ("Routing", "Routing Number","Text",    "Yes", ""),
    ("Routing", "Routing Name",  "Text",    "Yes", ""),
    ("Routing", "Comment",       "Text",    "No",  ""),
    ("Other Charges", "Serial Number",       "Integer", "Yes", ""),
    ("Other Charges", "FG Item ID",          "Text",    "Yes", ""),
    ("Other Charges", "BOM Number",          "Text",    "No",  ""),
    ("Other Charges", "Labour Cost",         "Float",   "Yes", ""),
    ("Other Charges", "Labour Comment",      "Text",    "No",  ""),
    ("Other Charges", "Machinery Cost",      "Float",   "Yes", ""),
    ("Other Charges", "Machinery Comment",   "Text",    "No",  ""),
    ("Other Charges", "Electricity Cost",    "Float",   "Yes", ""),
    ("Other Charges", "Electricity Comment", "Text",    "No",  ""),
    ("Other Charges", "Other Cost",          "Float",   "Yes", ""),
    ("Other Charges", "Other Comment",       "Text",    "No",  ""),
    ("Other Charges", "CF1",                 "Depends", "Depends", ""),
]


def parse_qty_unit(value):
    """Parse '1PCS' → (1, 'PCS'), '0.002KGS' → (0.002, 'KGS'), '1' → (1, 'PCS')."""
    if value is None:
        return 1, "PCS"
    s = str(value).strip()
    if not s:
        return 1, "PCS"
    m = re.match(r'^(\d+(?:\.\d+)?)\s*([A-Za-z]*)$', s)
    if m:
        try:
            qty = float(m.group(1))
            qty = int(qty) if qty == int(qty) else qty
        except ValueError:
            qty = 1
        unit = m.group(2).upper().strip() or "PCS"
        return qty, unit
    try:
        qty = float(s)
        return (int(qty) if qty == int(qty) else qty), "PCS"
    except (ValueError, TypeError):
        return 1, "PCS"


def _parse_tally_bom_xls(file_bytes):
    """XLS branch for parse_tally_bom — uses xlrd to read bold/italic font info."""
    wb  = xlrd.open_workbook(file_contents=file_bytes, formatting_info=True)
    snames = wb.sheet_names()
    ws  = wb.sheet_by_name("Item Estimates") if "Item Estimates" in snames else wb.sheet_by_index(0)

    part_col   = 0
    qty_col    = 1
    data_start = 6  # 0-indexed fallback

    for row_idx in range(min(15, ws.nrows)):
        row = ws.row(row_idx)
        for ci, cell in enumerate(row):
            if cell.value and str(cell.value).strip().lower() == "particulars":
                part_col   = ci
                data_start = row_idx + 1
                for cj, hcell in enumerate(row):
                    if hcell.value and str(hcell.value).strip().lower() in ("qty", "quantity"):
                        qty_col = cj
                break
        if data_start != 6:
            break

    fg_rows, rm_rows = [], []
    fg_sl         = 0
    parent        = None
    rm_seq        = 0
    parent_fg_qty = 1

    for row_idx in range(data_start, ws.nrows):
        row = ws.row(row_idx)
        if part_col >= len(row):
            continue
        cell_part = row[part_col]
        name = str(cell_part.value).strip() if cell_part.value not in (None, "", xlrd.empty_cell.value) else ""
        if not name:
            continue

        xf      = wb.xf_list[cell_part.xf_index]
        font    = wb.font_list[xf.font_index]
        bold    = bool(font.bold)
        italic  = bool(font.italic)
        qty_raw = row[qty_col].value if qty_col < len(row) else None

        if bold and not italic:
            fg_sl += 1
            fg_qty, uom = parse_qty_unit(qty_raw)
            fg_rows.append({
                "Sl_No": fg_sl, "FG Item Name": name,
                "FG UOM": uom, "BOM Name": name, "FG Cost Allocation": 100,
            })
            parent        = fg_sl
            parent_fg_qty = fg_qty
            rm_seq        = 0
        elif italic:
            if parent is None:
                continue
            rm_seq += 1
            qty, unit = parse_qty_unit(qty_raw)
            if parent_fg_qty and parent_fg_qty > 1:
                qty = qty / parent_fg_qty
            rm_rows.append({
                "Sl_No": parent, "#": rm_seq,
                "Item Description": name, "Quantity": qty, "Unit": unit,
            })

    return fg_rows, rm_rows


def parse_tally_bom(file_bytes):
    """
    Parse Tally BOM Excel (Item Estimates sheet).

    Supports two Tally export layouts automatically:
      Layout A (e.g. Velvu):  Col A = Particulars, Col B = Qty   — header ~row 5
      Layout B (e.g. TZ):     Col A = Item Code,   Col B = Particulars, Col C = Qty — header ~row 8

    Detection: scan first 15 rows for a cell whose value is "Particulars".
    That cell's column index becomes part_col; the adjacent "Qty" column becomes qty_col.
    Font is always read from the Particulars column cell.

    Font rules:
      Bold only       → FG  (new parent)
      Italic only     → RM of current parent
      Bold + Italic   → SFG: added as RM of current parent (parent context unchanged)

    Supports both .xlsx and .xls formats.
    Returns (fg_rows, rm_rows) as lists of dicts.
    """
    if file_bytes[:4] == b'\xd0\xcf\x11\xe0':   # XLS magic bytes
        return _parse_tally_bom_xls(file_bytes)

    # ── XLSX path (openpyxl) ──────────────────────────────────────────
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb["Item Estimates"] if "Item Estimates" in wb.sheetnames else wb.active

    # ── Auto-detect header row & column positions ─────────────────────
    part_col   = 0   # index of Particulars column (item name)
    qty_col    = 1   # index of Qty column
    data_start = 6   # fallback: row after header

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15), start=1):
        for ci, cell in enumerate(row):
            if cell.value and str(cell.value).strip().lower() == "particulars":
                part_col   = ci
                data_start = row_idx + 1
                # Find Qty in the same header row
                for cj, hcell in enumerate(row):
                    if hcell.value and str(hcell.value).strip().lower() in ("qty", "quantity"):
                        qty_col = cj
                break
        if data_start != 6:
            break

    # ── Parse data rows ───────────────────────────────────────────────
    fg_rows, rm_rows = [], []
    fg_sl         = 0
    parent        = None
    rm_seq        = 0
    parent_fg_qty = 1

    for row in ws.iter_rows(min_row=data_start):
        if part_col >= len(row):
            continue
        cell_part = row[part_col]
        cell_qty  = row[qty_col] if qty_col < len(row) else None

        name = str(cell_part.value).strip() if cell_part.value not in (None, "") else ""
        if not name:
            continue

        bold    = bool(cell_part.font and cell_part.font.bold)
        italic  = bool(cell_part.font and cell_part.font.italic)
        qty_raw = cell_qty.value if cell_qty else None

        if bold and not italic:
            # ── FG ──────────────────────────────────────────────────────
            fg_sl += 1
            fg_qty, uom = parse_qty_unit(qty_raw)
            fg_rows.append({
                "Sl_No": fg_sl, "FG Item Name": name,
                "FG UOM": uom, "BOM Name": name, "FG Cost Allocation": 100,
            })
            parent        = fg_sl
            parent_fg_qty = fg_qty
            rm_seq        = 0

        elif italic:
            # ── RM or SFG-as-RM ─────────────────────────────────────────
            if parent is None:
                continue
            rm_seq += 1
            qty, unit = parse_qty_unit(qty_raw)
            if parent_fg_qty and parent_fg_qty > 1:
                qty = qty / parent_fg_qty
            rm_rows.append({
                "Sl_No": parent, "#": rm_seq,
                "Item Description": name, "Quantity": qty, "Unit": unit,
            })

    return fg_rows, rm_rows


def make_bom_xlsx(fg_rows, rm_rows):
    """Produce BulkUpload-format Excel (FG + RM + Scrap + Routing + Other Charges + Instructions)."""
    from openpyxl.styles import PatternFill, Font as XLFont
    RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    RED_FONT = XLFont(color="FFFFFF", bold=True)

    name_count = {}
    for r in fg_rows:
        n = r["FG Item Name"]
        name_count[n] = name_count.get(n, 0) + 1
    duplicates = {n for n, c in name_count.items() if c > 1}

    wb    = Workbook()

    # ── FG sheet ─────────────────────────────────────────────────────────
    ws_fg = wb.active
    ws_fg.title = "FG"
    ws_fg.append(BOM_FG_HEADERS)
    for r in fg_rows:
        ws_fg.append([
            r["Sl_No"], None, r["FG Item Name"], r["FG UOM"],
            None, r["BOM Name"], None, None, None, None,
            r["FG Cost Allocation"], None, None,
        ])
        if r["FG Item Name"] in duplicates:
            rn = ws_fg.max_row
            for c in range(1, len(BOM_FG_HEADERS) + 1):
                ws_fg.cell(row=rn, column=c).fill = RED_FILL
                ws_fg.cell(row=rn, column=c).font = RED_FONT

    # ── RM sheet ─────────────────────────────────────────────────────────
    ws_rm = wb.create_sheet("RM")
    ws_rm.append(BOM_RM_HEADERS)
    for r in rm_rows:
        ws_rm.append([
            r["Sl_No"], None, None, r["#"], None,
            r["Item Description"], r["Quantity"], r["Unit"], None,
        ])

    # ── Scrap sheet ──────────────────────────────────────────────────────
    ws_scrap = wb.create_sheet("Scrap")
    ws_scrap.append(BOM_SCRAP_HEADERS)

    # ── Routing sheet ────────────────────────────────────────────────────
    ws_routing = wb.create_sheet("Routing")
    ws_routing.append(BOM_ROUTING_HEADERS)

    # ── Other Charges sheet ──────────────────────────────────────────────
    ws_other = wb.create_sheet("Other Charges")
    ws_other.append(BOM_OTHER_CHARGES_HEADERS)

    # ── Instructions sheet ───────────────────────────────────────────────
    ws_instr = wb.create_sheet("Instructions")
    ws_instr.append(BOM_INSTRUCTIONS_HEADERS)
    for row in BOM_INSTRUCTIONS_ROWS:
        ws_instr.append(list(row))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def bom_filename(fname):
    m  = re.search(r"\(([^)]+)\)", fname)
    cn = m.group(1) if m else fname.rsplit(".", 1)[0]
    return f"BOM_Upload_({cn}).xlsx"


# ─────────────────────────────────────────────────────────────────────────────
# AI BOM converter helpers  (Other Formats — Gemini powered)
# ─────────────────────────────────────────────────────────────────────────────

def get_bom_preview(file_bytes):
    """
    Return (preview_text_for_gemini, rows_for_display).
    Shows up to 25 non-empty rows with actual Excel row numbers and column letters.
    Supports xlsx, xls, and csv.
    """
    is_xlsx = file_bytes[:4] == b'PK\x03\x04'
    if is_xlsx:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
        ws = wb.active
        rows_with_idx = []
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True), start=1):
            if any(v is not None for v in row):
                rows_with_idx.append((idx, list(row)))
            if len(rows_with_idx) >= 25:
                break
    else:
        # XLS or CSV — use read_file
        sheets = read_file(file_bytes)
        sname  = pick_sheet(sheets)
        all_rows = sheets[sname]
        rows_with_idx = [
            (i + 1, list(r)) for i, r in enumerate(all_rows)
            if any(v is not None for v in r)
        ][:25]

    if not rows_with_idx:
        return "Empty file.", []

    max_cols = max(len(r) for _, r in rows_with_idx)
    padded   = [(i, r + [None] * (max_cols - len(r))) for i, r in rows_with_idx]
    col_labels = [chr(65 + i) for i in range(min(max_cols, 26))]

    lines = ["Excel Row | " + " | ".join(col_labels)]
    for idx, row in padded:
        vals = [(str(v)[:25] if v is not None else "(blank)") for v in row]
        lines.append(f"Row {idx:3d}   | " + " | ".join(vals))

    display_rows = [r for _, r in padded]
    return "\n".join(lines), display_rows


def call_gemini_bom(api_key, chat_history, preview_text, fname):
    """Send chat history to Gemini and return the response text."""
    from google import genai
    from google.genai import types

    sys_prompt = (
        f"You are a BOM (Bill of Materials) conversion assistant for TranZact.\n\n"
        f"The user uploaded a file named '{fname}'. Here is a preview:\n\n"
        f"{preview_text}\n\n"
        "Columns are labelled A, B, C... (A = index 0, B = index 1, ...).\n"
        "Rows are shown as actual Excel row numbers.\n\n"
        "Your goal: understand how the file identifies FG (Finished Good) vs RM (Raw Material) "
        "rows, then immediately output a JSON conversion spec.\n\n"
        "The user will describe the BOM logic in their own words — any format, any language. "
        "Your job is to map their description to the right spec. "
        "If the user's message has enough info, output the spec IMMEDIATELY. "
        "Only ask ONE follow-up if something truly critical is missing.\n\n"
        "Supported hierarchy methods and their spec shapes:\n\n"
        "1. COLUMN VALUE — a column contains FG/RM/SFG or similar labels:\n"
        '{"ready":true,"header_row":<n>,"item_name_col":<n>,"qty_col":<n>,'
        '"hierarchy_method":"column_value","hierarchy_col":<n>,'
        '"fg_values":["FG"],"rm_values":["RM"],"sfg_values":["SFG"]}\n\n'
        "2. FONT STYLE — bold=FG, italic=RM, bold+italic=SFG:\n"
        '{"ready":true,"header_row":<n>,"item_name_col":<n>,"qty_col":<n>,'
        '"hierarchy_method":"font"}\n\n'
        "3. INDENTATION — leading spaces distinguish FG from RM:\n"
        '{"ready":true,"header_row":<n>,"item_name_col":<n>,"qty_col":<n>,'
        '"hierarchy_method":"indentation","fg_indent":0,"rm_indent":4}\n\n'
        "4. LEVEL NUMBER — a column has level numbers like 1, 2, 3 (or L1, L2) where 1=FG, 2+=RM:\n"
        '{"ready":true,"header_row":<n>,"item_name_col":<n>,"qty_col":<n>,'
        '"hierarchy_method":"level","level_col":<n>,'
        '"fg_levels":[1,"1","L1"],"rm_levels":[2,3,"2","3","L2","L3"]}\n\n'
        "5. HIERARCHICAL NUMBERING — item number column has 1, 1.1, 1.1.1 etc. "
        "(top-level = FG, sub-levels = RM):\n"
        '{"ready":true,"header_row":<n>,"item_name_col":<n>,"qty_col":<n>,'
        '"hierarchy_method":"numbering","number_col":<n>}\n\n'
        "6. PRODUCTION TYPE — sub-parts have a Production Type column; "
        "BOUGHT OUT=direct RM, SUB CONTRACT=SFG (sent to party, received back):\n"
        '{"ready":true,"header_row":<n>,"item_name_col":<n>,"qty_col":<n>,'
        '"hierarchy_method":"production_type","fg_identifier":"font",'
        '"production_type_col":<n>,'
        '"bought_out_values":["BOUGHT OUT"],"sub_contract_values":["SUB CONTRACT"]}\n\n'
        "Pick the method that best matches the user's description. "
        "If still unclear, ask exactly ONE brief question."
    )

    import time
    client = genai.Client(api_key=api_key)

    # Build history from all but the last message
    history = []
    for msg in chat_history[:-1]:
        role = "user" if msg["role"] == "user" else "model"
        history.append(types.Content(
            role=role,
            parts=[types.Part(text=msg["content"])]
        ))

    chat = client.chats.create(
        model="gemini-2.5-flash",
        config=types.GenerateContentConfig(system_instruction=sys_prompt),
        history=history
    )

    # Retry up to 3 times on 503 temporary overload
    for attempt in range(3):
        try:
            response = chat.send_message(chat_history[-1]["content"])
            return response.text
        except Exception as e:
            if "503" in str(e) and attempt < 2:
                time.sleep(4 * (attempt + 1))   # 4s, 8s
                continue
            raise


def extract_bom_spec(text):
    """Extract JSON conversion spec from Gemini response text."""
    import json
    # Try ```json ... ``` block
    m = re.search(r'```json\s*(.*?)\s*```', text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(1))
        except Exception:
            pass
    # Try first JSON object in the text
    m = re.search(r'\{.*\}', text, re.DOTALL)
    if m:
        try:
            return json.loads(m.group(0))
        except Exception:
            pass
    return None


def _to_xlsx_bytes(file_bytes):
    """If file_bytes is CSV or XLS, convert to in-memory xlsx for openpyxl."""
    is_xlsx = file_bytes[:4] == b'PK\x03\x04'
    if is_xlsx:
        return file_bytes
    sheets = read_file(file_bytes)
    sname  = pick_sheet(sheets)
    rows   = sheets[sname]
    wb2    = Workbook()
    ws2    = wb2.active
    for row in rows:
        ws2.append([v for v in row])
    buf = io.BytesIO()
    wb2.save(buf)
    return buf.getvalue()


def apply_bom_spec(file_bytes, spec):
    """
    Apply a Gemini-generated conversion spec to a BOM file.
    Returns (fg_rows, rm_rows) in the same format as parse_tally_bom.
    """
    file_bytes = _to_xlsx_bytes(file_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    header_row  = spec.get("header_row", 0)     # 0-indexed → openpyxl row = header_row+1
    item_col    = spec.get("item_name_col", 0)
    qty_col_idx = spec.get("qty_col", 1)
    method      = spec.get("hierarchy_method", "column_value")
    data_start  = header_row + 2                 # first data row in openpyxl (1-indexed)

    fg_rows, rm_rows = [], []
    fg_sl         = 0
    parent        = None
    rm_seq        = 0
    parent_fg_qty = 1

    for row in ws.iter_rows(min_row=data_start):
        if item_col >= len(row):
            continue
        cell_name = row[item_col]
        name = str(cell_name.value).strip() if cell_name.value not in (None, "") else ""
        if not name:
            continue

        cell_qty = row[qty_col_idx] if qty_col_idx < len(row) else None
        qty_raw  = cell_qty.value if cell_qty else None
        is_fg = is_rm = False

        if method == "column_value":
            hier_col = spec.get("hierarchy_col", 0)
            hcell    = row[hier_col] if hier_col < len(row) else None
            hval     = str(hcell.value).strip().upper() if (hcell and hcell.value) else ""
            fg_vals  = [v.upper() for v in spec.get("fg_values",  ["FG"])]
            rm_vals  = [v.upper() for v in spec.get("rm_values",  ["RM"])] + \
                       [v.upper() for v in spec.get("sfg_values", ["SFG"])]
            is_fg = hval in fg_vals
            is_rm = hval in rm_vals

        elif method == "indentation":
            fg_ind = spec.get("fg_indent", 0)
            rm_ind = spec.get("rm_indent", 2)
            raw    = str(cell_name.value) if cell_name.value else ""
            leading = len(raw) - len(raw.lstrip())
            is_fg = leading == fg_ind
            is_rm = leading >= rm_ind

        elif method == "font":
            bold   = bool(cell_name.font and cell_name.font.bold)
            italic = bool(cell_name.font and cell_name.font.italic)
            is_fg  = bold and not italic
            is_rm  = italic

        elif method == "level":
            lv_col  = spec.get("level_col", 0)
            lv_cell = row[lv_col] if lv_col < len(row) else None
            lv_val  = str(lv_cell.value).strip() if (lv_cell and lv_cell.value is not None) else ""
            fg_lvls = [str(v) for v in spec.get("fg_levels", ["1"])]
            rm_lvls = [str(v) for v in spec.get("rm_levels", ["2", "3"])]
            try:
                lv_num = str(int(float(lv_val)))
            except (ValueError, TypeError):
                lv_num = lv_val
            is_fg = lv_num in fg_lvls or lv_val in fg_lvls
            is_rm = lv_num in rm_lvls or lv_val in rm_lvls

        elif method == "numbering":
            num_col  = spec.get("number_col", 0)
            num_cell = row[num_col] if num_col < len(row) else None
            num_val  = str(num_cell.value).strip() if (num_cell and num_cell.value is not None) else ""
            # top-level = single integer like "1", "2" → FG; sub-level has dots "1.1" → RM
            parts = num_val.split(".")
            is_fg = len(parts) == 1 and parts[0].isdigit()
            is_rm = len(parts) > 1

        elif method == "production_type":
            # FG identification uses fg_identifier (font/indent/column_value)
            fg_id = spec.get("fg_identifier", "font")
            if fg_id == "font":
                bold  = bool(cell_name.font and cell_name.font.bold)
                italic = bool(cell_name.font and cell_name.font.italic)
                is_fg = bold and not italic
            elif fg_id == "indentation":
                raw = str(cell_name.value) if cell_name.value else ""
                is_fg = (len(raw) - len(raw.lstrip())) == spec.get("fg_indent", 0)
            elif fg_id == "column_value":
                hcol  = spec.get("hierarchy_col", 0)
                hcell = row[hcol] if hcol < len(row) else None
                hval  = str(hcell.value).strip().upper() if (hcell and hcell.value) else ""
                is_fg = hval in [v.upper() for v in spec.get("fg_values", ["FG"])]

            if not is_fg:
                pt_col  = spec.get("production_type_col", -1)
                pt_cell = row[pt_col] if (pt_col >= 0 and pt_col < len(row)) else None
                pt_val  = str(pt_cell.value).strip().upper() if (pt_cell and pt_cell.value) else ""
                bo_vals  = [v.upper() for v in spec.get("bought_out_values",  ["BOUGHT OUT"])]
                sc_vals  = [v.upper() for v in spec.get("sub_contract_values", ["SUB CONTRACT"])]

                if pt_val in bo_vals:
                    is_rm = True          # direct RM — purchased as-is
                elif pt_val in sc_vals:
                    # SFG: add as RM under parent FG AND as its own FG entry
                    if parent is not None:
                        rm_seq += 1
                        qty, unit = parse_qty_unit(qty_raw)
                        if parent_fg_qty and parent_fg_qty > 1:
                            qty = qty / parent_fg_qty
                        rm_rows.append({
                            "Sl_No": parent, "#": rm_seq,
                            "Item Description": name, "Quantity": qty, "Unit": unit,
                        })
                    fg_sl += 1
                    sfg_qty, uom = parse_qty_unit(qty_raw)
                    fg_rows.append({
                        "Sl_No": fg_sl, "FG Item Name": name,
                        "FG UOM": uom, "BOM Name": name, "FG Cost Allocation": 100,
                    })
                    parent        = fg_sl
                    parent_fg_qty = sfg_qty
                    rm_seq        = 0
                    continue   # already handled, skip generic is_fg/is_rm block

        if is_fg:
            fg_sl += 1
            fg_qty, uom = parse_qty_unit(qty_raw)
            fg_rows.append({
                "Sl_No": fg_sl, "FG Item Name": name,
                "FG UOM": uom, "BOM Name": name, "FG Cost Allocation": 100,
            })
            parent        = fg_sl
            parent_fg_qty = fg_qty
            rm_seq        = 0
        elif is_rm and parent is not None:
            rm_seq += 1
            qty, unit = parse_qty_unit(qty_raw)
            if parent_fg_qty and parent_fg_qty > 1:
                qty = qty / parent_fg_qty
            rm_rows.append({
                "Sl_No": parent, "#": rm_seq,
                "Item Description": name, "Quantity": qty, "Unit": unit,
            })

    return fg_rows, rm_rows


def read_sheet_rows(file_bytes, sheet_name=None):
    sheets = read_file(file_bytes)
    if sheet_name:
        rows = sheets.get(sheet_name)
        if rows is None:
            raise KeyError(sheet_name)
    else:
        rows = sheets[pick_sheet(sheets)]
    return rows, list(sheets.keys())

# ─────────────────────────────────────────────────────────────────────────────
# Page config & CSS
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="TranZact · Master Data Converter",
    page_icon="⚡",
    layout="centered"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;600;700;900&display=swap');

/* ── Logo bar ── */
.logo-bar {
    display: flex;
    align-items: center;
    gap: 12px;
    padding: 18px 0 6px 0;
}

/* ── Header subtitle bar ── */
.header-bar {
    background: linear-gradient(135deg, #1a1b17 0%, #2a1f2d 100%);
    border: 1px solid #292a27;
    border-left: 4px solid #da5d37;
    border-radius: 10px;
    padding: 14px 20px;
    margin: 4px 0 12px 0;
    display: flex;
    align-items: center;
    gap: 18px;
    flex-wrap: wrap;
}
.header-bar .pill {
    background: #232420;
    border: 1px solid #3a3b37;
    border-radius: 20px;
    padding: 3px 12px;
    font-size: 0.78rem;
    color: #aaaca6;
    white-space: nowrap;
}

/* ── Upload zone ── */
[data-testid="stFileUploader"] {
    border: 2px dashed #da5d37 !important;
    border-radius: 12px !important;
    padding: 6px !important;
}

/* ── Expander cards ── */
[data-testid="stExpander"] {
    border: 1px solid #2e2f2b !important;
    border-radius: 10px !important;
    background: #141510 !important;
}

/* ── Mapping preview grid ── */
.map-grid {
    display: grid;
    grid-template-columns: 1fr auto 1fr;
    gap: 6px 10px;
    align-items: center;
    margin: 8px 0;
}
.map-src  { background:#1e2a1e; color:#7ecf7e; border-radius:6px; padding:4px 10px; font-size:0.82rem; font-family:monospace; }
.map-tgt  { background:#1e1e2a; color:#7e9ecf; border-radius:6px; padding:4px 10px; font-size:0.82rem; }
.map-arr  { color:#555; font-size:0.9rem; text-align:center; }
.map-extra{ background:#2a2117; color:#c8a96e; border-radius:6px; padding:3px 9px; font-size:0.78rem; font-family:monospace; display:inline-block; margin:2px; }

/* ── Network stat cards ── */
.stat-cards { display:flex; gap:12px; margin:14px 0; }
.stat-card  {
    flex:1; border-radius:12px; padding:16px 18px;
    display:flex; flex-direction:column; gap:4px;
}
.stat-card .sc-num  { font-size:2rem; font-weight:800; font-family:'Outfit',sans-serif; line-height:1; }
.stat-card .sc-lbl  { font-size:0.78rem; font-weight:600; letter-spacing:0.03em; opacity:0.8; }
.stat-green { background:#0f2318; border:1px solid #1f4a30; }
.stat-green .sc-num { color:#4ade80; }
.stat-green .sc-lbl { color:#86efac; }
.stat-amber { background:#231a0a; border:1px solid #4a3010; }
.stat-amber .sc-num { color:#fbbf24; }
.stat-amber .sc-lbl { color:#fcd34d; }
.stat-red   { background:#1f0e0e; border:1px solid #4a1f1f; }
.stat-red   .sc-num { color:#f87171; }
.stat-red   .sc-lbl { color:#fca5a5; }

/* ── Format badge ── */
.fmt-badge {
    display:inline-block; border-radius:20px; padding:2px 12px;
    font-size:0.75rem; font-weight:700; letter-spacing:0.04em;
    background:#1e2630; color:#7eb8f0; border:1px solid #2d4a6a;
    margin-bottom:10px;
}

/* ── Template card ── */
.tmpl-card {
    background:#141510; border:1px solid #2e2f2b;
    border-radius:10px; padding:14px 18px; margin-bottom:10px;
}

/* ── Divider ── */
hr { border-color: #292a27 !important; }

/* ── Tab labels ── */
button[data-baseweb="tab"] { font-weight: 600 !important; font-size:0.9rem !important; }
</style>
""", unsafe_allow_html=True)

# ── Header ──────────────────────────────────────────────────────────────────
st.markdown("""
<div class="logo-bar">
  <svg width="40" height="40" viewBox="0 0 42 42" style="flex-shrink:0">
    <line x1="21" y1="21" x2="4"  y2="38" stroke="#F5A623" stroke-width="9" stroke-linecap="round"/>
    <line x1="21" y1="21" x2="38" y2="4"  stroke="#5CB85C" stroke-width="9" stroke-linecap="round"/>
    <line x1="21" y1="21" x2="4"  y2="4"  stroke="#4A90D9" stroke-width="9" stroke-linecap="round"/>
    <line x1="21" y1="21" x2="38" y2="38" stroke="#E05A3C" stroke-width="9" stroke-linecap="round"/>
    <circle cx="21" cy="21" r="5" fill="#0e0f0c"/>
  </svg>
  <span style="font-family:'Outfit',sans-serif;font-size:1.72rem;font-weight:900;
               letter-spacing:-0.5px;color:#fff;line-height:1;">TRANZACT</span>
  <span style="display:inline-flex;align-items:center;gap:3px;
               border:2px solid #E07A5F;border-radius:7px;padding:3px 9px 3px 7px;
               font-family:'Outfit',sans-serif;font-size:0.84rem;font-weight:700;
               color:#fff;line-height:1;margin-left:2px;">
    Ai&nbsp;<span style="color:#F5D76E;font-size:0.73rem;vertical-align:middle;">✦</span>
  </span>
</div>
<div class="header-bar">
  <span style="color:#aaaca6;font-size:0.88rem;">Master Data Converter</span>
  <span class="pill">📦 Item Master</span>
  <span class="pill">🏢 Network Master</span>
  <span class="pill">🧩 BOM Upload</span>
  <span class="pill">⚡ Auto-detects any format</span>
</div>
""", unsafe_allow_html=True)

st.divider()
tab1, tab2, tab3, tab4 = st.tabs(["📦  Item Master", "🏢  Network Master", "🧩  BOM Upload", "🗂  Templates"])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 : Any client format  →  Product_Add_(X).xlsx
# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    st.markdown("Upload any client item file — columns are **auto-detected and mapped** to your format.")

    up = st.file_uploader("Upload client file", type=["xlsx", "xls", "csv"], key="t1_up",
                          label_visibility="collapsed")

    if up:
        up.seek(0)
        fbytes = up.read()
        fname  = up.name

        if st.session_state.get("t1_fname") != fname:
            sheets     = read_file(fbytes)
            sname      = pick_item_master_sheet(sheets)
            rows       = sheets[sname]
            hidx, hdrs = detect_header(rows)
            data_rows  = [r for r in rows[hidx + 1:] if any(v for v in r)]
            templates  = load_templates()
            tmpl_name, tmpl = find_template(hdrs, templates)

            if tmpl:
                mapping    = tmpl["mapping"].copy()
                extra_cols = tmpl.get("extra_cols", [])
            else:
                mapping, extra_cols = auto_map(hdrs, data_rows)

            st.session_state.t1_fname   = fname
            st.session_state.t1_rows    = rows
            st.session_state.t1_hidx    = hidx
            st.session_state.t1_headers = hdrs
            st.session_state.t1_tname   = tmpl_name
            st.session_state.t1_mapping = mapping
            st.session_state.t1_extra   = extra_cols
            st.session_state.t1_out     = None

        rows       = st.session_state.t1_rows
        hidx       = st.session_state.t1_hidx
        hdrs       = st.session_state.t1_headers
        tname      = st.session_state.t1_tname
        mapping    = st.session_state.t1_mapping
        extra_cols = st.session_state.t1_extra
        data_rows  = [r for r in rows[hidx + 1:] if any(v for v in r)]

        # ── File info strip ──────────────────────────────────────────────
        col_a, col_b, col_c = st.columns(3)
        col_a.metric("File", fname.rsplit(".", 1)[0][:22])
        col_b.metric("Rows detected", len(data_rows))
        col_c.metric("Template", tname if tname else "Auto-mapped")

        # ── Mapping preview ──────────────────────────────────────────────
        with st.expander("🔍  Detected column mapping", expanded=False):
            if mapping:
                rows_html = ""
                for tgt, src in mapping.items():
                    if src:
                        rows_html += f'<div class="map-grid"><span class="map-src">{src}</span><span class="map-arr">→</span><span class="map-tgt">{tgt}</span></div>'
                if extra_cols:
                    extras_html = "".join(f'<span class="map-extra">{e}</span>' for e in extra_cols[:12])
                    rows_html += f'<div style="margin-top:10px;"><span style="color:#888;font-size:0.8rem;">Extra columns appended: </span>{extras_html}</div>'
                st.markdown(rows_html, unsafe_allow_html=True)
            else:
                st.caption("No mapping detected yet.")

        st.markdown("")
        if st.button("▶  Convert Now", type="primary", use_container_width=True, key="t1_go"):
            with st.spinner("Converting…"):
                result = do_convert(rows, hidx, mapping, extra_cols)
            if result:
                st.session_state.t1_out      = make_xlsx(result, extra_cols)
                st.session_state.t1_out_name = out_filename(fname)
                st.session_state.t1_out_cnt  = len(result)
            else:
                st.error("❌  No items found — could not identify the Item Name column.")

        if st.session_state.get("t1_out") and st.session_state.get("t1_fname") == fname:
            cnt = st.session_state.t1_out_cnt
            st.success(f"✅  **{cnt} items** converted successfully!")
            st.download_button(
                "⬇️  Download Product_Add File",
                data=st.session_state.t1_out,
                file_name=st.session_state.t1_out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="t1_dl"
            )

            if not tname:
                with st.expander("💾  Save mapping as template for this client"):
                    sc1, sc2 = st.columns([3, 1])
                    tname_in = sc1.text_input(
                        "Template name:",
                        placeholder="e.g.  MSHRIY  /  BOM Client  /  ITEM",
                        key="t1_tname_in",
                        label_visibility="collapsed"
                    )
                    if sc2.button("Save", key="t1_save_btn", use_container_width=True):
                        if tname_in.strip():
                            all_t = load_templates()
                            all_t[tname_in.strip()] = {
                                "fingerprint": sorted(h for h in hdrs if h),
                                "mapping":     {k: v for k, v in mapping.items() if v},
                                "extra_cols":  extra_cols
                            }
                            save_templates(all_t)
                            st.session_state.t1_tname = tname_in.strip()
                            st.success(f"✅  Saved as **{tname_in.strip()}**!")

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 : Network Master  →  Network_Add_(X).xlsx
# ─────────────────────────────────────────────────────────────────────────────
with tab2:
    st.markdown("Upload a client ledger / vendor file — auto-converts to your **Network Add** format.")

    net_up = st.file_uploader("Upload client file", type=["xlsx", "xls", "csv"], key="net_up",
                               label_visibility="collapsed")

    if net_up:
        net_up.seek(0)
        net_bytes = net_up.read()
        net_fname = net_up.name

        if st.session_state.get("net_fname") != net_fname:
            try:
                sheets   = read_file(net_bytes)
                im_flag, im_cols = is_item_master_file(sheets)
                if im_flag:
                    st.error(
                        f"❌ This looks like an **Item Master** file "
                        f"(detected columns: *{', '.join(sorted(im_cols))}*). "
                        f"Please upload it in the **📦 Item Master** tab instead."
                    )
                    st.stop()
                detected = []
                for sname, srows in sheets.items():
                    fmt, hidx = detect_network_format(srows)
                    if fmt != "unknown":
                        data_cnt = sum(1 for r in srows[hidx + 1:] if any(v for v in r if v))
                        if data_cnt > 0:
                            detected.append({"name": sname, "rows": srows,
                                             "fmt": fmt, "hidx": hidx, "count": data_cnt})
                # Fallback: nothing auto-detected — try most-likely sheet
                if not detected:
                    sname     = _tally_detect_sheet(sheets)
                    srows     = sheets[sname]
                    fmt, hidx = detect_network_format(srows)
                    data_cnt  = sum(1 for r in srows[hidx + 1:] if any(v for v in r if v))
                    detected  = [{"name": sname, "rows": srows,
                                  "fmt": fmt, "hidx": hidx, "count": data_cnt}]
                st.session_state.net_fname    = net_fname
                st.session_state.net_detected = detected
                st.session_state.net_out      = None
            except Exception as e:
                st.error(f"❌  Could not read file: {e}")
                st.stop()

        detected   = st.session_state.net_detected
        fmt_labels = {"tally": "Tally Export", "mshriy": "MSHRIY Format",
                      "generic": "Generic Format", "unknown": "Unknown"}

        # Show one badge per detected sheet
        badges = " ".join(
            f'<span class="fmt-badge">⚙ {fmt_labels.get(d["fmt"], d["fmt"])} · {d["name"]}</span>'
            for d in detected
        )
        st.markdown(badges, unsafe_allow_html=True)

        col_a, col_b = st.columns(2)
        col_a.metric("File", net_fname.rsplit(".", 1)[0][:28])
        col_b.metric("Rows detected", sum(d["count"] for d in detected))
        if len(detected) > 1:
            st.caption("📋 Party data found in: " +
                       ", ".join(f'**{d["name"]}** ({d["count"]} rows)' for d in detected))

        any_known = any(d["fmt"] != "unknown" for d in detected)
        if not any_known:
            st.warning("⚠️  Could not detect file format. Please check the file.")
        else:
            st.markdown("")
            if st.button("▶  Convert Now", type="primary", use_container_width=True, key="net_go"):
                with st.spinner("Converting…"):
                    try:
                        all_parties = []
                        for d in detected:
                            if d["fmt"] == "tally":
                                ps = convert_tally_parties(d["rows"], d["hidx"])
                            elif d["fmt"] == "mshriy":
                                ps = convert_mshriy_parties(d["rows"], d["hidx"])
                            elif d["fmt"] == "generic":
                                ps = convert_generic_parties(d["rows"], d["hidx"])
                            else:
                                continue
                            all_parties.extend(ps)
                        parties = apply_pincode_lookup(all_parties, load_pincode_db())
                        parties = fill_addr2_with_city(parties)
                        ready, have_gstin, manual, duplicates = split_network_sheets(parties)
                        st.session_state.net_out      = make_network_xlsx(ready, have_gstin, manual, duplicates)
                        st.session_state.net_out_name = net_filename(net_fname)
                        st.session_state.net_counts   = (len(ready), len(have_gstin), len(manual), len(duplicates), len(parties))
                    except Exception as e:
                        st.error(f"❌  Something went wrong: {e}")

        if st.session_state.get("net_out") and st.session_state.get("net_fname") == net_fname:
            r, g, m, d, total = st.session_state.net_counts
            st.markdown(f"""
            <div class="stat-cards">
              <div class="stat-card stat-green">
                <div class="sc-num">{r}</div>
                <div class="sc-lbl">✅ READY TO UPLOAD</div>
              </div>
              <div class="stat-card stat-amber">
                <div class="sc-num">{g}</div>
                <div class="sc-lbl">🔑 HAVE GSTIN</div>
              </div>
              <div class="stat-card stat-red">
                <div class="sc-num">{m}</div>
                <div class="sc-lbl">✏️ NEED MANUAL UPDATE</div>
              </div>
              <div class="stat-card" style="background:#1a1025;border:1px solid #3d2060;">
                <div class="sc-num" style="color:#c084fc;">{d}</div>
                <div class="sc-lbl" style="color:#d8b4fe;">🔁 DUPLICATE GSTINs</div>
              </div>
            </div>
            <p style="color:#666;font-size:0.78rem;margin:0 0 12px 0;">{total} parties processed total &nbsp;·&nbsp; 🔴 Red rows = non-standard party type — review with client &nbsp;·&nbsp; 🔴 Red PIN cell = incomplete pincode</p>
            """, unsafe_allow_html=True)

            st.download_button(
                "⬇️  Download Network Add File",
                data=st.session_state.net_out,
                file_name=st.session_state.net_out_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
                key="net_dl"
            )

# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 : BOM Upload  →  BOM_Upload_(X).xlsx
# ─────────────────────────────────────────────────────────────────────────────
with tab3:
    st.markdown("Convert client BOM files to the **BulkUpload** format.")

    bom_section = st.radio(
        "Select BOM format",
        ["📊  Tally BOM", "📋  Other Formats"],
        horizontal=True,
        key="bom_section",
        label_visibility="collapsed"
    )

    st.divider()

    if bom_section == "📊  Tally BOM":
        st.markdown(
            "Upload a Tally BOM export — **bold** rows = FG, "
            "**italic** rows = RM, **bold+italic** rows = SFG (treated as RM under its parent)."
        )

        bom_up = st.file_uploader(
            "Upload Tally BOM file", type=["xlsx", "xls", "csv"],
            key="bom_up", label_visibility="collapsed"
        )

        if bom_up and bom_up.name.lower().endswith(".csv"):
            st.warning("⚠️  CSV files don't contain font formatting — FG/RM detection relies on **bold/italic** fonts and won't work. Please upload an **.xlsx** or **.xls** file for best results.")

        if bom_up:
            bom_up.seek(0)
            bom_bytes = bom_up.read()
            bom_fname = bom_up.name

            if st.session_state.get("bom_fname") != bom_fname:
                try:
                    fg_rows, rm_rows = parse_tally_bom(bom_bytes)
                    st.session_state.bom_fname   = bom_fname
                    st.session_state.bom_fg_rows = fg_rows
                    st.session_state.bom_rm_rows = rm_rows
                    st.session_state.bom_out     = None
                except Exception as e:
                    st.error(f"❌  Could not parse BOM file: {e}")
                    st.stop()

            fg_rows = st.session_state.bom_fg_rows
            rm_rows = st.session_state.bom_rm_rows

            name_cnt = {}
            for r in fg_rows:
                n = r["FG Item Name"]
                name_cnt[n] = name_cnt.get(n, 0) + 1
            dup_count = sum(1 for c in name_cnt.values() if c > 1)

            c1, c2, c3 = st.columns(3)
            c1.metric("FG Items",      len(fg_rows))
            c2.metric("RM Entries",    len(rm_rows))
            c3.metric("Duplicate FGs", dup_count)

            if dup_count:
                st.warning(
                    f"⚠️  **{dup_count} duplicate FG name(s)** found — "
                    "highlighted in red in the output file. Review before uploading."
                )

            st.markdown("")
            if st.button("▶  Convert Now", type="primary", use_container_width=True, key="bom_go"):
                with st.spinner("Converting…"):
                    bom_out = make_bom_xlsx(fg_rows, rm_rows)
                    st.session_state.bom_out      = bom_out
                    st.session_state.bom_out_name = bom_filename(bom_fname)

            if st.session_state.get("bom_out") and st.session_state.get("bom_fname") == bom_fname:
                st.success(
                    f"✅  **{len(fg_rows)} FG items** and **{len(rm_rows)} RM entries** converted!"
                )
                st.download_button(
                    "⬇️  Download BOM Upload File",
                    data=st.session_state.bom_out,
                    file_name=st.session_state.bom_out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                    key="bom_dl"
                )

    else:
        st.markdown(
            "Upload any BOM file and **describe the logic in plain language** — "
            "the AI assistant will ask clarifying questions, then convert it to the BulkUpload format."
        )

        # ── Gemini API key ────────────────────────────────────────────────
        raw_key = st.secrets.get("GEMINI_API_KEY", "") if hasattr(st, "secrets") else ""
        if not raw_key:
            raw_key = st.text_input(
                "Gemini API Key", type="password", key="gemini_key",
                placeholder="Paste your Gemini API key",
                label_visibility="collapsed"
            )
        api_key = raw_key.strip() if raw_key else ""

        if not api_key:
            st.info("Enter your Gemini API key above to get started.")
        else:
            other_up = st.file_uploader(
                "Upload BOM file (any format)", type=["xlsx", "xls", "csv"],
                key="other_bom_up", label_visibility="collapsed"
            )

            if other_up:
                other_up.seek(0)
                other_bytes = other_up.read()
                other_fname = other_up.name

                # Reset state on new file upload
                if st.session_state.get("other_bom_fname") != other_fname:
                    prev_text, prev_disp = get_bom_preview(other_bytes)
                    st.session_state.other_bom_fname  = other_fname
                    st.session_state.other_prev_text  = prev_text
                    st.session_state.other_prev_disp  = prev_disp
                    st.session_state.other_chat       = []
                    st.session_state.other_spec       = None
                    st.session_state.other_bom_out    = None

                # ── File preview ──────────────────────────────────────────
                with st.expander("📋  File preview", expanded=True):
                    disp = st.session_state.other_prev_disp
                    if disp:
                        max_c = max(len(r) for r in disp)
                        cols  = [chr(65 + i) for i in range(min(max_c, 26))]
                        table = [
                            {"Row": i + 1, **{cols[j]: (v if v is not None else "") for j, v in enumerate(r)}}
                            for i, r in enumerate(disp)
                        ]
                        st.dataframe(table, use_container_width=True, hide_index=True)

                # ── Chat history ──────────────────────────────────────────
                for msg in st.session_state.other_chat:
                    with st.chat_message(msg["role"]):
                        st.markdown(msg["content"])

                # ── Ready to convert ──────────────────────────────────────
                if st.session_state.other_spec:
                    st.success("✅  Logic understood — ready to convert.")
                    if st.button("▶  Convert Now", type="primary",
                                 use_container_width=True, key="other_go"):
                        with st.spinner("Converting…"):
                            try:
                                fg_r, rm_r = apply_bom_spec(
                                    other_bytes, st.session_state.other_spec
                                )
                                st.session_state.other_bom_out      = make_bom_xlsx(fg_r, rm_r)
                                st.session_state.other_bom_out_name = bom_filename(other_fname)
                                st.session_state.other_counts       = (len(fg_r), len(rm_r))
                            except Exception as e:
                                st.error(f"❌  Conversion failed: {e}")

                    if st.session_state.get("other_bom_out"):
                        fg_c, rm_c = st.session_state.other_counts
                        st.success(f"✅  **{fg_c} FG items** and **{rm_c} RM entries** converted!")
                        st.download_button(
                            "⬇️  Download BOM Upload File",
                            data=st.session_state.other_bom_out,
                            file_name=st.session_state.other_bom_out_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary", use_container_width=True, key="other_dl"
                        )

                # ── Chat input (shown until spec is ready) ────────────────
                else:
                    prompt = st.chat_input(
                        "Explain the BOM logic — e.g. 'Col B = item name, Col C = qty, bold rows are FG, italic are RM, header is row 8'",
                        key="other_chat_input"
                    )
                    if prompt:
                        st.session_state.other_chat.append(
                            {"role": "user", "content": prompt}
                        )
                        with st.spinner("Thinking…"):
                            try:
                                reply = call_gemini_bom(
                                    api_key,
                                    st.session_state.other_chat,
                                    st.session_state.other_prev_text,
                                    other_fname
                                )
                            except Exception as e:
                                reply = f"⚠️  API error: {e}"

                        spec = extract_bom_spec(reply)
                        if spec and spec.get("ready"):
                            st.session_state.other_spec = spec
                            display = (
                                "Got it! I fully understand the format. "
                                "Click **▶ Convert Now** to proceed."
                            )
                        else:
                            display = reply

                        st.session_state.other_chat.append(
                            {"role": "assistant", "content": display}
                        )
                        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 : Templates
# ─────────────────────────────────────────────────────────────────────────────
with tab4:
    templates = load_templates()

    if not templates:
        st.markdown("""
        <div style="text-align:center;padding:40px 20px;color:#666;">
          <div style="font-size:2.5rem;margin-bottom:12px;">🗂</div>
          <div style="font-size:1rem;font-weight:600;color:#888;margin-bottom:6px;">No templates saved yet</div>
          <div style="font-size:0.84rem;">Convert a client file in <b>Item Master</b> tab and click <b>💾 Save mapping</b> — next time the same client's file is auto-recognised.</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown(f"**{len(templates)} saved template{'s' if len(templates) != 1 else ''}** — uploaded when a matching client file is detected automatically.")
        st.markdown("")
        for name, tmpl in list(templates.items()):
            with st.expander(f"📋  {name}"):
                m  = tmpl.get("mapping", {})
                ec = tmpl.get("extra_cols", [])
                fp = tmpl.get("fingerprint", [])

                rows_html = ""
                for t, c in m.items():
                    if c:
                        rows_html += f'<div class="map-grid"><span class="map-src">{c}</span><span class="map-arr">→</span><span class="map-tgt">{t}</span></div>'
                if ec:
                    extras = "".join(f'<span class="map-extra">{e}</span>' for e in ec[:10])
                    rows_html += f'<div style="margin-top:8px;"><span style="color:#888;font-size:0.78rem;">Extra: </span>{extras}</div>'
                if rows_html:
                    st.markdown(rows_html, unsafe_allow_html=True)
                if fp:
                    preview = ", ".join(fp[:6]) + ("…" if len(fp) > 6 else "")
                    st.caption(f"Matched by {len(fp)} columns · {preview}")
                st.markdown("")
                if st.button(f"🗑  Delete  '{name}'", key=f"del_{name}"):
                    del templates[name]
                    save_templates(templates)
                    st.rerun()

st.divider()
st.markdown(
    '<p style="text-align:center;color:#444;font-size:0.78rem;margin:0;">'
    'TranZact Ai &nbsp;·&nbsp; Master Data Converter &nbsp;·&nbsp; v3.0 &nbsp;·&nbsp; '
    'Item Master &nbsp;·&nbsp; Network Master &nbsp;·&nbsp; BOM Upload</p>',
    unsafe_allow_html=True
)
